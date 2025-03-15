from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
from selenium.webdriver.chrome.options import Options
from datetime import datetime

def capturar_dados_clima(cidade):
    options = Options()
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
    navegador = webdriver.Chrome(options=options)
    navegador.get('https://www.climaeradar.com.br/previsao-tempo/sao-paulo/3705599')
    try:
        temperatura_elemento = WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/wo-root/wo-app-shell/div/div/wo-weather/wo-weather-header/div/div[1]/wo-nowcast-card/div[3]/wo-nowcast-conditions-left/span'))
        )
        temperatura = temperatura_elemento.text

        umidade_elemento = WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/wo-root/wo-app-shell/div/div/wo-weather/wo-weather-header/div/div[1]/wo-hourcast-card/div[1]/wo-horizontal-scroll-arrows/div/div[1]/wo-forecast-hour[1]/wo-if-visible/wo-weather-characteristics-precipitation/div')) # Substitua pelo XPATH da umidade
        )
        umidade = umidade_elemento.text

        time.sleep(5)
        navegador.quit()
        return [temperatura, umidade]  
    except:
        navegador.quit()
        return None  


def salvar_dados_em_xlsx(dados, nome_arquivo="clima.xlsx"):
    try:
        arquivo = load_workbook(nome_arquivo)
        plan = arquivo['lista']
    except FileNotFoundError:
        from openpyxl import Workbook
        workbook = Workbook()
        plan = workbook.active
        plan.title = 'lista'
        plan["A1"] = "Temperatura (°C)"
        plan["B1"] = "Umidade (%)"
        plan["C1"] = "Data e Hora"
        arquivo = workbook

    data_hora_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    dados.append(data_hora_atual)
    plan.append(dados)
    arquivo.save(nome_arquivo)